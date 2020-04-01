import arcpy
import pyodbc
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

fonts = [Font(color=colors.DARKBLUE),Font(color=colors.DARKYELLOW),Font(color="ff6500")] 


base = "c:/fsh/"
base += "sde/"
#fgdbs = ['fecc.gdb','yaml.gdb','ucs.gdb','bart.gdb']
fgdbs = ['fecc.sde','yam.sde','ucs.sde','bart.sde','joe.sde','jas.sde']
axinfcs = ['DES_','Fielding_','RDL_']
fieldnames = ['name','aliasName','baseName','defaultValue','domain','editable','isNullable','length','precision','required','scale','type']
workbook = Workbook()
sheet = workbook.active
sheet.title = "Databases"
cols = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

all_fcs_and_tables = []
geodb_fcs = {}
all_dommains = []
geodb_domains = {}

colcount = 0
for fgdb in fgdbs:

    if fgdb not in geodb_fcs:
        geodb_fcs[fgdb] = []
    if fgdb not in geodb_domains:
        geodb_domains[fgdb] = []

    arcpy.env.workspace = base+fgdb
    newsheet = workbook.create_sheet(f"FeatureClassesFields {fgdb[:-4]}")
    print(base+fgdb)
    sheet[f"{cols[colcount]}1"] = f"Feature Classes {base+fgdb}"
    featureclasses = arcpy.ListFeatureClasses()
    rowcount = 2
    subrowcount = 1
    subcolcount = 0
    for fc in sorted(featureclasses):
        fc = fc.split(".")[-1]
        if fc not in all_fcs_and_tables:
            all_fcs_and_tables.append(fc)
        
        if fc not in geodb_fcs[fgdb]:
            geodb_fcs[fgdb].append(fc)

        sheet[f"{cols[colcount]}{rowcount}"] = fc
        rowcount += 1
        fdfc = [fc]# for x in axinfcs if x.upper() in fc.upper()]
        if fdfc:
            newsheet[f"{cols[subcolcount]}{subrowcount}"].font = fonts[0]
            newsheet[f"{cols[subcolcount]}{subrowcount}"] = fdfc[0]
            subrowcount += 1
            desc = arcpy.Describe(fdfc[0])
            fd_header = False
            for fd in  desc.fields:
                if not fd_header:
                    for k,an_attr in enumerate(fieldnames):
                        newsheet[f"{cols[k]}{subrowcount}"] = an_attr
                    fd_header = True
                    subrowcount += 1 
                for k,an_attr in enumerate(fieldnames):
                    newsheet[f"{cols[k]}{subrowcount}"] = getattr(fd,an_attr)
                
                subrowcount += 1

    colcount+=1
    rowcount = 2
    for tbl in arcpy.ListTables():
        tbl = tbl.split(".")[-1]
        if tbl not in all_fcs_and_tables:
            all_fcs_and_tables.append(tbl)
        
        if tbl not in geodb_fcs[fgdb]:
            geodb_fcs[fgdb].append(tbl)

        sheet[f"{cols[colcount]}{rowcount}"] = tbl
        rowcount += 1
    
    for a_domain in arcpy.da.ListDomains(base+fgdb):
        if a_domain.name not in all_dommains:
            all_dommains.append(a_domain.name)
        
        if a_domain.name not in geodb_fcs[fgdb]:
            geodb_domains[fgdb].append(a_domain.name)

    colcount+=1

compsheet = workbook.create_sheet(f"Compare")
compcol = 0
comprow = 1
fontidx = 0
fcfdrowidx = 1

for fgdb in fgdbs:
    arcpy.env.workspace = base+fgdb
    compsheet[f"{cols[compcol]}{comprow}"] = fgdb
    comprow += 1
    featureclasses = arcpy.ListFeatureClasses()
    sorted_cs = []
    for afc in featureclasses:
        afc = afc.split(".")[-1]
        sorted_cs = sorted_cs + [afc for x in axinfcs if x.upper() in afc.upper()]
    for fcname in sorted(sorted_cs):
        if fcname not in workbook:
            fcwksht = workbook.create_sheet(fcname)
        else:
            fcwksht = workbook[fcname]
            
        subsheetfontidx = compcol - len(fonts) if compcol >= len(fonts) else compcol

        fcwksht[f"{cols[compcol]}{fcfdrowidx}"].font = fonts[subsheetfontidx]
        fcwksht[f"{cols[compcol]}{fcfdrowidx}"] = fgdb
        fcfdrowidx += 1
        desc = arcpy.Describe(fcname)
        compsheet[f"{cols[compcol]}{comprow}"].font = fonts[fontidx] 
        compsheet[f"{cols[compcol]}{comprow}"] = fcname

        fcwksht[f"{cols[compcol]}{fcfdrowidx}"].font = fonts[subsheetfontidx] 
        fcwksht[f"{cols[compcol]}{fcfdrowidx}"] = fcname
        fcfdrowidx += 1
        comprow += 1
        sorted_fields = sorted([fd.name for fd in  desc.fields])
        for srtfd in  sorted_fields:
            compsheet[f"{cols[compcol]}{comprow}"].font = fonts[fontidx]
            compsheet[f"{cols[compcol]}{comprow}"] = srtfd

            fcwksht[f"{cols[compcol]}{fcfdrowidx}"].font = fonts[subsheetfontidx]
            fcwksht[f"{cols[compcol]}{fcfdrowidx}"] = srtfd
            comprow += 1
            fcfdrowidx += 1
        fontidx = fontidx + 1 if fontidx != 2 else 0
        fcfdrowidx = 1
    compcol += 1
    comprow = 1


    subtrow = 1
    subtcol = 0
    subtvcol = 1
    nextcol = 0
    for fcname in sorted(sorted_cs):
        if f"SubTypes-{fgdb}" not in workbook:
            sbtwksht = workbook.create_sheet(f"SubTypes-{fgdb}")
        else:
            sbtwksht = workbook[f"SubTypes-{fgdb}"]
        subtypes = arcpy.da.ListSubtypes(f"{base+fgdb}/{fcname}")
        if subtypes:
            if [subtypevalue for subtypevalue,subtypedict in subtypes.items() if subtypevalue != 0]:
                sbtwksht[f"{cols[subtcol]}{subtrow}"].font = fonts[0]
                sbtwksht[f"{cols[subtcol]}{subtrow}"] = fcname
                subtrow += 1
        
                for subtypevalue,subtypedict in subtypes.items():

                    sbtwksht[f"{cols[subtcol]}{subtrow}"] = "Id"
                    sbtwksht[f"{cols[subtvcol]}{subtrow}"] = subtypevalue
                    subtrow +=1
                    sbtwksht[f"{cols[subtcol]}{subtrow}"] = "Name"
                    sbtwksht[f"{cols[subtvcol]}{subtrow}"].font = Font(color=colors.RED)
                    sbtwksht[f"{cols[subtvcol]}{subtrow}"] = subtypedict["Name"]
                    subtrow +=1
                    sbtwksht[f"{cols[subtcol]}{subtrow}"] = "SubtypeField"
                    sbtwksht[f"{cols[subtvcol]}{subtrow}"] = subtypedict["SubtypeField"]
                    subtrow +=1
                    for fn,fv in subtypedict['FieldValues'].items():
                        if fv[0] is not None or fv[1] is not None:
                            sbtwksht[f"{cols[subtcol]}{subtrow}"] = f"Field:{fn}"
                            sbtwksht[f"{cols[subtvcol]}{subtrow}"] =fv[0]
                            sbtwksht[f"{cols[subtvcol+1]}{subtrow}"] =fv[1].name
                            subtrow +=1

        print(subtypes)
        


# final sheet - work out all the fcs and which are in what
allsheet = workbook.create_sheet(f"AllFCsToDBs")
dommainsheet = workbook.create_sheet(f"Domains")

allcol = 0
all_row = 1

allsheet[f"{cols[allcol]}{all_row}"] = "Table"
allcol += 1

col_lookup = {}

for k,v in enumerate(geodb_fcs.keys()):
    allsheet[f"{cols[allcol]}{all_row}"] = v
    col_lookup[v] = allcol
    allcol += 1


allcol = 0
for k,v in enumerate(sorted(all_fcs_and_tables)):
    # k is our row
    allsheet[f"{cols[allcol]}{k+2}"] = v
    
    for db,col in col_lookup.items():
        if v in geodb_fcs[db]:
            allsheet[f"{cols[col]}{k+2}"].font = fonts[0]
            allsheet[f"{cols[col]}{k+2}"] = "Y"
        else:
            allsheet[f"{cols[col]}{k+2}"].font = fonts[2]
            allsheet[f"{cols[col]}{k+2}"] = "N"

allcol = 0
all_row = 1
dommainsheet[f"{cols[allcol]}{all_row}"] = "Domain"
allcol += 1

col_lookup = {}
for k,v in enumerate(geodb_domains.keys()):
    dommainsheet[f"{cols[allcol]}{all_row}"] = v
    col_lookup[v] = allcol
    allcol += 1

allcol = 0
for k,v in enumerate(sorted(all_dommains)):
    # k is our row
    dommainsheet[f"{cols[allcol]}{k+2}"] = v
    for db,col in col_lookup.items():
        if v in geodb_domains[db]:
            dommainsheet[f"{cols[col]}{k+2}"].font = fonts[0]
            dommainsheet[f"{cols[col]}{k+2}"] = "Y"
        else:
            dommainsheet[f"{cols[col]}{k+2}"].font = fonts[2]
            dommainsheet[f"{cols[col]}{k+2}"] = "N"

for fgdb in fgdbs:
    gdbappdomainsheet = workbook.create_sheet(f"AppliedDomains-{fgdb}")
    appdomrow = 1
    gdbappdomainsheet[f"{cols[0]}{appdomrow}"] = "Class Name"
    gdbappdomainsheet[f"{cols[1]}{appdomrow}"] = "Field Name"
    gdbappdomainsheet[f"{cols[2]}{appdomrow}"] = "Subtype Name"
    appdomrow +=1
    db = fgdb.split(".")[0]
    conn_string = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER=nrtc1sql1.database.windows.net;DATABASE=nrtc1{db}1;UID=gisadmin;PWD=Ow5UHy80QxxHH3v3tmQl"
    conn = pyodbc.connect(conn_string) #, @param_out = @out OUTPUT;
    cursor = conn.cursor()
       
    
    for a_domain in arcpy.da.ListDomains(base+fgdb):
        try:
            # Class Name	Field Name	Subtype Name
            # nrtc1ucs1.GISADMIN.DES_cable	cable_size	Distribution UG
            cursor.execute("{CALL ShowAppliedDomains(?)}",a_domain.name)
            rows = cursor.fetchall()
            while rows:
                for row in rows:
                    gdbappdomainsheet[f"{cols[0]}{appdomrow}"] = "Null" if row[0] is None else row[0]
                    gdbappdomainsheet[f"{cols[1]}{appdomrow}"] = "Null" if row[1] is None else row[1]
                    gdbappdomainsheet[f"{cols[2]}{appdomrow}"] = "Null" if row[2] is None else row[2]
                    appdomrow += 1
                if cursor.nextset():
                    rows = cursor.fetchall()
                else:
                    rows = None
        except Exception as e:
            pass
    cursor.close()
    conn.close()
    del cursor
    del conn
    appdomrow = 1
    


for fgdb in fgdbs:
    gdbdomainsheet = workbook.create_sheet(f"Domains-{fgdb}")
    domains = arcpy.da.ListDomains(base+fgdb)
    dom_row = 1
    for domain in domains:
        gdbdomainsheet[f"{cols[0]}{dom_row}"].font = fonts[0]
        gdbdomainsheet[f"{cols[0]}{dom_row}"] = domain.name
        dom_row += 1
        if domain.domainType == 'CodedValue':
            coded_values = domain.codedValues
            for val, desc in coded_values.items():
                gdbdomainsheet[f"{cols[0]}{dom_row}"] = val
                gdbdomainsheet[f"{cols[1]}{dom_row}"] = desc
                dom_row += 1
        elif domain.domainType == 'Range':
            gdbdomainsheet[f"{cols[0]}{dom_row}"] = 'Min: {0}'.format(domain.range[0])
            gdbdomainsheet[f"{cols[1]}{dom_row}"] = 'Max: {0}'.format(domain.range[1])
            dom_row += 1


    def print_out_rule(featclass,attwksheet,att_row,rules):
        for ar in rules:
            if "Calculation" in ar.type: 
                att_col = 0
                attwksheet[f"{cols[att_col]}{att_row}"] = f"Calculation Rule:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = featclass
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Name:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.name
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Creation time:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.creationTime
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Field:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.fieldName
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Subtype code:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.subtypeCode
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Description:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.description
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Is editable:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.userEditable
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Is enabled:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.isEnabled
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Evaluation order:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.evaluationOrder
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Exclude from client evaluation:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.excludeFromClientEvaluation
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Triggering events:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =str(ar.triggeringEvents)
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Script expression:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.scriptExpression
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Is flagged as a batch rule:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.batch
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Severity:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.severity
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = "Tags:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.tags
                att_col -= 1
                att_row += 1
            elif "Constraint" in ar.type:  
                att_col = 0     
                attwksheet[f"{cols[att_col]}{att_row}"] = "Constraint Rule:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = featclass
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Name:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.name
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Creation time:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.creationTime
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Subtype code:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.subtypeCode
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Description:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.description
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Is editable:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.userEditable
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Is enabled:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.isEnabled
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Error number:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.errorNumber
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Error message:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.errorMessage
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Exclude from client evaluation:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.excludeFromClientEvaluation
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Triggering events:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.triggeringEvents
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Script expression:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.scriptExpression
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Tags:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = ar.tags
                att_col -= 1
                att_row += 1

            elif "Validation" in ar.type: 
                att_col = 0      
                attwksheet[f"{cols[att_col]}{att_row}"] =" Validation Rule:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] = featclass
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =" Name:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.name
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =" Creation time:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.creationTime
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Subtype code:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.subtypeCode
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Description:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.description
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Is enabled:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.isEnabled
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Error number:"
                att_col += 1
                fattwksheet[f"{cols[att_col]}{att_row}"] =ar.errorNumber
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =" Error message:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.errorMessage
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Script expression:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.scriptExpression
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Is flagged as a batch rule:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.batch
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Severity:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.severity
                att_col -= 1
                att_row += 1
                attwksheet[f"{cols[att_col]}{att_row}"] ="Tags:"
                att_col += 1
                attwksheet[f"{cols[att_col]}{att_row}"] =ar.tags
                att_col -= 1
                att_row += 1
            att_row += 1
        return att_row

    arcpy.env.workspace = base+fgdb
    attrulesheet = workbook.create_sheet(f"AttRules-{fgdb}")
    att_row = 1
    for afc in arcpy.ListFeatureClasses():
        desc = arcpy.Describe(afc)
        attRules = arcpy.Describe(afc).attributeRules
        if attRules:
            a_fc = afc.split(".")[-1]
            att_row = print_out_rule(a_fc,attrulesheet,att_row,attRules)
        #
    for tbl in arcpy.ListTables():
        desc = arcpy.Describe(tbl)
        attRules = arcpy.Describe(tbl).attributeRules
        if attRules:
            a_tbl = tbl.split(".")[-1]
            att_row = print_out_rule(a_tbl, attrulesheet,att_row,attRules)
        #


workbook.save(filename=f"dbs{datetime.now().strftime('%Y%m%d%H%M')}.xlsx")
